import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os

def gerar_relatorio_exportacao(df_filtrado):
    """
    Gera o relatório Excel formatado a partir do DataFrame filtrado.
    Retorna o caminho do arquivo temporário gerado.
    """
    colunas_saida = ["MÁQUINA", "DATA", "MATERIAL", "BLOCO", "DEMANDA", "PROCESSO", "OBSERVAÇÃO DE PRODUÇÃO", "QTD. CHAPAS", "VOLUME M²", "FEITO?"]
    
    df_out = pd.DataFrame()
    df_out["MÁQUINA"] = df_filtrado["SETOR"]
    
    def format_date(d):
        if pd.isna(d) or str(d).strip() == "" or str(d) == "nan": return "Sem Data"
        try:
            if isinstance(d, pd.Timestamp): return d.strftime("%d/%m/%Y")
            if "/" in str(d): return pd.to_datetime(str(d), format="%d/%m/%Y").strftime("%d/%m/%Y")
            return pd.to_datetime(str(d)).strftime("%d/%m/%Y")
        except: return str(d)
        
    df_out["DATA"] = df_filtrado["DATA"].apply(format_date)
    df_out["MATERIAL"] = df_filtrado["MATERIAL"]
    df_out["BLOCO"] = df_filtrado["BLOCO"]
    df_out["DEMANDA"] = df_filtrado["DEMANDA"]
    df_out["PROCESSO"] = df_filtrado["PROCESSO"]
    df_out["OBSERVAÇÃO DE PRODUÇÃO"] = df_filtrado["OBSERVAÇÃO DE PRODUÇÃO"]
    df_out["QTD. CHAPAS"] = pd.to_numeric(df_filtrado["QTD. CHAPAS"], errors="coerce").fillna(0)
    df_out["VOLUME M²"] = pd.to_numeric(df_filtrado["VOLUME M²"], errors="coerce").fillna(0)
    df_out["FEITO?"] = ""
    
    # Remove registros que não têm data definida, se desejar (ou mantém como 'Sem Data')
    # Como é para chão de fábrica, normalmente filtra-se apenas os que têm data alvo.
    
    df_out = df_out.sort_values(by=["MÁQUINA", "DATA"])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROGRAMAÇÃO GERAL"
    
    green_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Verde/Azul parecido com a foto
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    font_white_bold = Font(color="FFFFFF", bold=True)
    font_bold = Font(bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A1:J2")
    cell_title = ws.cell(row=1, column=1, value="PROGRAMAÇÃO - GERAL")
    cell_title.font = Font(size=20, bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx, col_name in enumerate(colunas_saida, 1):
        c = ws.cell(row=3, column=col_idx, value=col_name)
        c.fill = green_fill
        c.font = font_white_bold
        c.alignment = align_center
        c.border = border_thin
        
    current_row = 4
    total_chapas_geral = 0
    total_vol_geral = 0
    
    for maquina, df_maq in df_out.groupby("MÁQUINA", sort=False):
        for data_val, df_data in df_maq.groupby("DATA", sort=False):
            subtotal_chapas = 0
            subtotal_vol = 0
            
            for _, row_data in df_data.iterrows():
                ws.cell(row=current_row, column=1, value=row_data["MÁQUINA"]).border = border_thin
                ws.cell(row=current_row, column=2, value=row_data["DATA"]).border = border_thin
                ws.cell(row=current_row, column=3, value=row_data["MATERIAL"]).border = border_thin
                ws.cell(row=current_row, column=4, value=row_data["BLOCO"]).border = border_thin
                ws.cell(row=current_row, column=5, value=row_data["DEMANDA"]).border = border_thin
                ws.cell(row=current_row, column=6, value=row_data["PROCESSO"]).border = border_thin
                ws.cell(row=current_row, column=7, value=row_data["OBSERVAÇÃO DE PRODUÇÃO"]).border = border_thin
                
                chapas = row_data["QTD. CHAPAS"]
                vol = row_data["VOLUME M²"]
                
                ws.cell(row=current_row, column=8, value=chapas).border = border_thin
                ws.cell(row=current_row, column=9, value=vol).border = border_thin
                ws.cell(row=current_row, column=10, value="").border = border_thin
                
                subtotal_chapas += chapas
                subtotal_vol += vol
                current_row += 1
                
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            sub_title = ws.cell(row=current_row, column=1, value=f"{data_val} Total")
            sub_title.font = font_bold
            sub_title.fill = grey_fill
            
            c_chapas = ws.cell(row=current_row, column=8, value=subtotal_chapas)
            c_chapas.font = font_bold
            c_chapas.fill = grey_fill
            
            c_vol = ws.cell(row=current_row, column=9, value=subtotal_vol)
            c_vol.font = font_bold
            c_vol.fill = grey_fill
            
            for c_idx in range(1, 11):
                ws.cell(row=current_row, column=c_idx).border = border_thin
                ws.cell(row=current_row, column=c_idx).fill = grey_fill
                
            total_chapas_geral += subtotal_chapas
            total_vol_geral += subtotal_vol
            current_row += 1
            
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    tot_title = ws.cell(row=current_row, column=1, value="Total Geral")
    tot_title.font = font_bold
    
    t_chapas = ws.cell(row=current_row, column=8, value=total_chapas_geral)
    t_chapas.font = font_bold
    
    t_vol = ws.cell(row=current_row, column=9, value=total_vol_geral)
    t_vol.font = font_bold
    
    for c_idx in range(1, 11):
        ws.cell(row=current_row, column=c_idx).border = border_thin
        
    larguras = [20, 15, 25, 10, 15, 25, 40, 15, 15, 10]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        
    tmp_path = os.path.join(tempfile.gettempdir(), "Programacao_Diaria.xlsx")
    wb.save(tmp_path)
    return tmp_path
