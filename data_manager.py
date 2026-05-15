# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
import os
import json

# ---------------------------------------------------------------------------
# CONFIGURAÇÃO DE CAMINHOS (lidos do config.json em runtime)
# ---------------------------------------------------------------------------
_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

_DEFAULT_CONFIG = {
    "DB_FILE": r"z:\PCP\PROJETOS MARLON\ProgramarProd\COSTA GRAN. - PROGRAMAÇÕES - BASE DE DADOS.xlsm",
    "APONTAMENTO_FILE": r"z:\PCP\PROJETOS MARLON\ProgramarProd\Apontamento Produção (REV 2).xlsx",
    # Nomes das abas (configuráveis pelo usuário)
    "SHEET_PROGRAMACAO": "DB",
    "SHEET_ENTREGUES": "ENTREGUES",
    "SHEET_BASE_DADOS": "BASE DE DADOS",
    "SHEET_AP_BD": "DB",
    "SHEET_AP_BASE": "BASE DADOS",
    "SHEET_AP_PARADAS": "PARADAS",
    "SHEET_AP_INSUMOS": "INSUMOS",
    "BLOCKS_FILE": r"z:\PCP\PROJETOS MARLON\ProgramarProd\PLANILHA BLOCOS.xlsb",
    "CHAPAS_FILE": r"z:\PCP\PROJETOS MARLON\ProgramarProd\Estoque Chapas 2026.xlsx"
}


def get_config():
    """Lê o config.json e retorna o dicionário de configurações."""
    try:
        if os.path.exists(_CONFIG_PATH):
            with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            for k, v in _DEFAULT_CONFIG.items():
                cfg.setdefault(k, v)
            return cfg
    except Exception as e:
        print(f"Erro ao ler config.json: {e}")
    return dict(_DEFAULT_CONFIG)


def save_config(cfg: dict):
    """Salva o dicionário de configurações no config.json."""
    try:
        with open(_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=4)
        return True
    except Exception as e:
        print(f"Erro ao salvar config.json: {e}")
        return False


def _get_db_file():
    return get_config()["DB_FILE"]


def _get_apontamento_file():
    return get_config()["APONTAMENTO_FILE"]


def _get_sheet(key: str) -> str:
    """Retorna o nome da aba configurada, com fallback para o default."""
    return get_config().get(key, _DEFAULT_CONFIG[key])


def get_sheet_names(filepath: str):
    """Retorna a lista de nomes de abas de um arquivo Excel."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        print(f"Erro ao ler abas de {filepath}: {e}")
        return []


# ---------------------------------------------------------------------------
# LEITURA DE DADOS
# ---------------------------------------------------------------------------

def normalize_bloco(bloco):
    """Padroniza o ID do bloco para string, removendo .0 e espaços."""
    if pd.isna(bloco) or str(bloco).strip() == "" or str(bloco).lower() == "nan":
        return ""
    b_str = str(bloco).strip().upper()
    if b_str.endswith(".0"):
        b_str = b_str[:-2]
    return b_str


def get_data():
    """Lê a aba de Programação e autocompleta o SETOR se estiver vazio usando a Base de Dados."""
    try:
        db_file = _get_db_file()
        sheet_prog = _get_sheet("SHEET_PROGRAMACAO")
        sheet_base = _get_sheet("SHEET_BASE_DADOS")
        
        # Lê os dados principais
        df = pd.read_excel(db_file, sheet_name=sheet_prog, engine="openpyxl")
        
        # --- LÓGICA DE AUTOCOMPLETAR SETOR (Self-Healing) ---
        try:
            df_base = pd.read_excel(db_file, sheet_name=sheet_base, engine="openpyxl")
            if "PROCESSO" in df_base.columns and "SETOR" in df_base.columns:
                # Função interna para limpar nomes de processo (remove espaços, acentos zoados e força upper)
                def clean_p(p):
                    return str(p).strip().upper().replace("", "A") # Ajuste básico para encoding
                
                # Cria mapa normalizado: {PROCESSO_LIMPO: SETOR_VALOR}
                df_base["P_LIMPO"] = df_base["PROCESSO"].apply(clean_p)
                mapa_setor = df_base.dropna(subset=["P_LIMPO", "SETOR"]).drop_duplicates("P_LIMPO").set_index("P_LIMPO")["SETOR"].to_dict()
                
                # Identifica onde o SETOR está vazio ou NaN
                mask_vazio = df["SETOR"].isna() | (df["SETOR"].astype(str).str.strip() == "") | (df["SETOR"].astype(str).str.lower() == "nan")
                
                # Preenche usando a versão limpa do processo atual
                df.loc[mask_vazio, "SETOR"] = df.loc[mask_vazio, "PROCESSO"].apply(clean_p).map(mapa_setor)
        except Exception as e_base:
            print(f"Aviso: Não foi possível carregar o mapa de setores da Base de Dados: {e_base}")
        # ----------------------------------------------------

        # Normalização crítica de BLOCO para evitar bugs de identificação
        if "BLOCO" in df.columns:
            df["BLOCO"] = df["BLOCO"].apply(normalize_bloco)
        return df
    except Exception as e:
        print(f"Erro ao carregar os dados: {e}")
        return pd.DataFrame()


def get_headers():
    """Lê os cabeçalhos diretamente com openpyxl para ter um mapa de colunas exato."""
    try:
        wb = openpyxl.load_workbook(_get_db_file(), read_only=True, data_only=True)
        ws = wb[_get_sheet("SHEET_PROGRAMACAO")]
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                # Normalização de cabeçalhos para evitar erros com 'VOLUME M²' e outros
                h_name = str(val).strip()
                headers[h_name] = col
                # Fallback para Volume M2 sem o símbolo especial se necessário
                if "VOLUME M" in h_name.upper():
                    headers["VOLUME M²"] = col
                    headers["VOLUME M2"] = col
        wb.close()
        return headers
    except Exception as e:
        print(f"Erro ao ler cabeçalhos: {e}")
        return {}


def get_base_dados():
    """Lê a aba de Base de Dados e retorna o DataFrame."""
    try:
        df = pd.read_excel(_get_db_file(), sheet_name=_get_sheet("SHEET_BASE_DADOS"), engine="openpyxl")
        return df
    except Exception as e:
        print(f"Erro ao carregar base de dados: {e}")
        return pd.DataFrame()


def get_data_entregues():
    """Lê a aba de Entregues para cálculo de capacidade."""
    try:
        df = pd.read_excel(_get_db_file(), sheet_name=_get_sheet("SHEET_ENTREGUES"), engine="openpyxl")
        return df
    except Exception as e:
        print(f"Erro ao carregar entregues: {e}")
        return pd.DataFrame()


# ---------------------------------------------------------------------------
# CÁLCULOS
# ---------------------------------------------------------------------------

def get_historico_medias(df):
    """
    Calcula a média histórica de QTD. CHAPAS por dia para cada PROCESSO.
    Considera apenas registros com STATUS PROCESSO == 'REALIZADO'.
    """
    try:
        realizados = df[df["STATUS PROCESSO"] == "REALIZADO"].copy()
        if realizados.empty:
            return {}

        def parse_data(val):
            if pd.isna(val) or val == "": return pd.NaT
            if isinstance(val, pd.Timestamp): return val
            try: return pd.to_datetime(str(val), format="%d/%m/%Y")
            except:
                try: return pd.to_datetime(str(val))
                except: return pd.NaT

        realizados["DATA_REALIZADA_DT"] = realizados["DATA REALIZADA"].apply(parse_data)
        realizados = realizados.dropna(subset=["DATA_REALIZADA_DT"])
        realizados["QTD. CHAPAS"] = pd.to_numeric(realizados["QTD. CHAPAS"], errors="coerce").fillna(0)

        producao_diaria = realizados.groupby(["PROCESSO", "DATA_REALIZADA_DT"])["QTD. CHAPAS"].sum().reset_index()
        medias = producao_diaria.groupby("PROCESSO")["QTD. CHAPAS"].mean().to_dict()
        return medias
    except Exception as e:
        print(f"Erro ao calcular médias: {e}")
        return {}


def get_historico_medias_entregues():
    """Calcula a média de capacidade usando a base histórica ENTREGUES."""
    df = get_data_entregues()
    if df.empty:
        return {}
    try:
        def parse_data(val):
            if pd.isna(val) or val == "": return pd.NaT
            if isinstance(val, pd.Timestamp): return val
            try: return pd.to_datetime(str(val), format="%d/%m/%Y")
            except:
                try: return pd.to_datetime(str(val))
                except: return pd.NaT

        col_data = "DATA REALIZADA" if "DATA REALIZADA" in df.columns else "DATA"
        df["DATA_REALIZADA_DT"] = df[col_data].apply(parse_data)
        df = df.dropna(subset=["DATA_REALIZADA_DT"])
        df["QTD. CHAPAS"] = pd.to_numeric(df.get("QTD. CHAPAS", 0), errors="coerce").fillna(0)

        producao_diaria = df.groupby(["SETOR", "DATA_REALIZADA_DT"])["QTD. CHAPAS"].sum().reset_index()
        medias = producao_diaria.groupby("SETOR")["QTD. CHAPAS"].mean().to_dict()
        return medias
    except Exception as e:
        print(f"Erro ao calcular médias de entregues: {e}")
        return {}


# ---------------------------------------------------------------------------
# VALIDAÇÕES DE NEGÓCIO
# ---------------------------------------------------------------------------

def validar_sequencia_bloco(df, bloco_id, index_atual, nova_data_str):
    """
    Valida se a etapa anterior do bloco foi finalizada ou tem data agendada
    anterior ou igual à nova_data.
    """
    try:
        if pd.isna(bloco_id) or str(bloco_id).strip() == "":
            return True, ""

        df_bloco = df[df["BLOCO"] == bloco_id].copy()
        df_bloco = df_bloco.sort_index()
        indices_bloco = df_bloco.index.tolist()

        if index_atual not in indices_bloco:
            return True, ""

        pos = indices_bloco.index(index_atual)

        if pos == 0:
            return True, ""

        index_anterior = indices_bloco[pos - 1]
        linha_anterior = df_bloco.loc[index_anterior]

        status_anterior = str(linha_anterior.get("STATUS PROCESSO", ""))
        if status_anterior == "REALIZADO":
            return True, ""

        data_anterior_str = linha_anterior.get("DATA", "")
        if pd.isna(data_anterior_str) or str(data_anterior_str).strip() == "":
            proc = linha_anterior.get("PROCESSO", "Desconhecido")
            return False, f"O processo anterior '{proc}' não foi agendado nem concluído."

        nova_data_dt = pd.to_datetime(nova_data_str, format="%d/%m/%Y")

        try:
            if isinstance(data_anterior_str, pd.Timestamp):
                data_anterior_dt = data_anterior_str
            else:
                data_anterior_str_clean = str(data_anterior_str).strip()
                if "/" in data_anterior_str_clean:
                    data_anterior_dt = pd.to_datetime(data_anterior_str_clean, format="%d/%m/%Y")
                else:
                    data_anterior_dt = pd.to_datetime(data_anterior_str_clean)

            if data_anterior_dt.date() > nova_data_dt.date():
                proc = linha_anterior.get("PROCESSO", "Desconhecido")
                return False, f"O processo anterior '{proc}' está agendado para {data_anterior_dt.strftime('%d/%m/%Y')}, que é depois da data escolhida."
        except Exception:
            pass  # Se não conseguir parsear, deixa passar

        return True, ""
    except Exception as e:
        print(f"Erro ao validar bloco: {e}")
        return False, "Erro interno na validação."


# ---------------------------------------------------------------------------
# ESCRITA DE DADOS
# ---------------------------------------------------------------------------

def add_record(record_dict):
    """Adiciona um novo registro preservando o formato e macros."""
    try:
        wb = openpyxl.load_workbook(_get_db_file(), keep_vba=True)
        ws = wb[_get_sheet("SHEET_PROGRAMACAO")]

        next_row = ws.max_row + 1
        for row in range(2, ws.max_row + 2):
            if ws.cell(row=row, column=1).value is None:
                next_row = row
                break

        headers = get_headers()

        for key, value in record_dict.items():
            if key in headers:
                col_idx = headers[key]
                ws.cell(row=next_row, column=col_idx, value=value)

        wb.save(_get_db_file())
        return True
    except Exception as e:
        print(f"Erro ao adicionar registro: {e}")
        return False


def add_records(records_list):
    """Adiciona múltiplos registros de uma vez (útil para roteiros de blocos)."""
    try:
        wb = openpyxl.load_workbook(_get_db_file(), keep_vba=True)
        ws = wb[_get_sheet("SHEET_PROGRAMACAO")]

        next_row = ws.max_row + 1
        for row in range(2, ws.max_row + 2):
            if ws.cell(row=row, column=1).value is None:
                next_row = row
                break

        headers = get_headers()

        for record_dict in records_list:
            for key, value in record_dict.items():
                if key in headers:
                    col_idx = headers[key]
                    ws.cell(row=next_row, column=col_idx, value=value)
            next_row += 1

        wb.save(_get_db_file())
        return True
    except Exception as e:
        print(f"Erro ao adicionar múltiplos registros: {e}")
        return False


def update_cell_by_row(df_index, updates_dict):
    """
    Atualiza células específicas baseando-se no índice do DataFrame.
    O índice do DataFrame 0 corresponde à linha 2 do Excel (linha 1 = cabeçalho).
    """
    try:
        wb = openpyxl.load_workbook(_get_db_file(), keep_vba=True)
        ws = wb[_get_sheet("SHEET_PROGRAMACAO")]

        excel_row = df_index + 2
        headers = get_headers()

        for col_name, new_value in updates_dict.items():
            if col_name in headers:
                col_idx = headers[col_name]
                # Se for None, limpa a célula explicitamente
                if new_value is None:
                    ws.cell(row=excel_row, column=col_idx).value = None
                else:
                    ws.cell(row=excel_row, column=col_idx, value=new_value)

        wb.save(_get_db_file())
        return True
    except Exception as e:
        print(f"Erro ao atualizar célula: {e}")
        return False


def update_base_dados(df_novo):
    """Sobrescreve as colunas PROCESSO e SETOR na aba de Base de Dados."""
    try:
        wb = openpyxl.load_workbook(_get_db_file(), keep_vba=True)
        ws = wb[_get_sheet("SHEET_BASE_DADOS")]

        col_proc = None
        col_setor = None

        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == "PROCESSO": col_proc = col
            elif val == "SETOR": col_setor = col

        if not col_proc or not col_setor:
            return False

        max_row = ws.max_row
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col_proc).value = None
            ws.cell(row=r, column=col_setor).value = None

        for idx, row in df_novo.iterrows():
            proc_val = str(row.get("PROCESSO", "")).strip()
            setor_val = str(row.get("SETOR", "")).strip()
            if proc_val and proc_val != "nan":
                ws.cell(row=idx + 2, column=col_proc, value=proc_val)
                ws.cell(row=idx + 2, column=col_setor, value=setor_val if setor_val != "nan" else "")

        wb.save(_get_db_file())
        return True
    except Exception as e:
        print(f"Erro ao atualizar base de dados: {e}")
        return False


def salvar_edicao_bloco_excel(bloco_id, material, demanda, qtd_chapas, vol_m2, roteiro_atual):
    """
    Sincroniza as edições de um bloco existente diretamente no Excel.
    Reescreve as linhas nas posições originais, adiciona novas linhas logo abaixo
    se o roteiro cresceu, ou deleta linhas finais se o roteiro diminuiu.
    """
    try:
        wb = openpyxl.load_workbook(_get_db_file(), keep_vba=True)
        ws = wb[_get_sheet("SHEET_PROGRAMACAO")]
        headers = get_headers()

        col_bloco = headers.get("BLOCO")
        old_rows = []
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=col_bloco).value).strip() == str(bloco_id).strip():
                old_rows.append(r)

        num_new = len(roteiro_atual)
        num_old = len(old_rows)
        last_written_row = max(old_rows) if old_rows else ws.max_row

        for i, step in enumerate(roteiro_atual):
            if i < num_old:
                row_to_write = old_rows[i]
            else:
                insert_idx = last_written_row + 1
                ws.insert_rows(insert_idx, 1)
                row_to_write = insert_idx

            last_written_row = row_to_write

            if headers.get("BLOCO"): ws.cell(row=row_to_write, column=headers["BLOCO"], value=bloco_id)
            if headers.get("MATERIAL"): ws.cell(row=row_to_write, column=headers["MATERIAL"], value=material)
            if headers.get("DEMANDA"): ws.cell(row=row_to_write, column=headers["DEMANDA"], value=demanda)
            if headers.get("QTD. CHAPAS"): ws.cell(row=row_to_write, column=headers["QTD. CHAPAS"], value=qtd_chapas)
            if headers.get("VOLUME M²"): ws.cell(row=row_to_write, column=headers["VOLUME M²"], value=vol_m2)

            if headers.get("PROCESSO"): ws.cell(row=row_to_write, column=headers["PROCESSO"], value=step.get("PROCESSO", ""))
            if headers.get("SETOR"): ws.cell(row=row_to_write, column=headers["SETOR"], value=step.get("SETOR", ""))
            if headers.get("OBSERVAÇÃO DE PRODUÇÃO"): ws.cell(row=row_to_write, column=headers["OBSERVAÇÃO DE PRODUÇÃO"], value=step.get("OBSERVACAO", ""))

            data_str = step.get("DATA", "")
            if data_str and headers.get("DATA"): ws.cell(row=row_to_write, column=headers["DATA"], value=data_str)

            data_realizada = step.get("DATA REALIZADA", "")
            if data_realizada and headers.get("DATA REALIZADA"): ws.cell(row=row_to_write, column=headers["DATA REALIZADA"], value=data_realizada)

        # Deletar as sobras físicas, de baixo para cima
        if num_old > num_new:
            sobras = old_rows[num_new:]
            for r in reversed(sobras):
                ws.delete_rows(r, 1)

        wb.save(_get_db_file())
        return True
    except Exception as e:
        print(f"Erro ao salvar edição do bloco no excel: {e}")
        return False


# ---------------------------------------------------------------------------
# APONTAMENTO DE PRODUÇÃO
# ---------------------------------------------------------------------------

def get_mapa_resumido_processos():
    """
    Lê a aba de mapeamento do arquivo de Apontamento e retorna um dicionário
    mapeando o nome completo do processo para o nome resumido.
    Ex: '19-POLIMENTO (S)' -> 'POLIMENTO'
    """
    try:
        df_bd = pd.read_excel(
            _get_apontamento_file(),
            sheet_name=_get_sheet("SHEET_AP_BASE"),
            header=0,
            engine="openpyxl",
            usecols=[0, 1]
        )
        df_bd.columns = ["PROCESSO_COMPLETO", "RESUMIDO"]
        df_bd = df_bd.dropna(subset=["PROCESSO_COMPLETO"])
        mapa = {}
        for _, row in df_bd.iterrows():
            proc = str(row["PROCESSO_COMPLETO"]).strip().upper()
            res = str(row["RESUMIDO"]).strip().upper() if pd.notna(row["RESUMIDO"]) else proc
            mapa[proc] = res
        return mapa
    except Exception as e:
        print(f"Erro ao ler BASE DADOS do Apontamento: {e}")
        return {}


def get_apontamentos_do_dia(data_alvo_date):
    """
    e retorna um DataFrame filtrado pela data_alvo_date.
    """
    try:
        # Busca dinâmica do cabeçalho (procura em até 100 linhas)
        df_full = pd.read_excel(_get_apontamento_file(), sheet_name=_get_sheet("SHEET_AP_BD"), header=None, engine="openpyxl", nrows=100)
        # Lista de todos os aliases para detecção rápida
        todos_aliases = ["DATA REG", "DATA", "DATA_REG", "DATA DO APONTAMENTO", "DATA LANÇAMENTO", "LANÇAMENTO", "DATA CADASTRO", "DATA_APONTAMENTO", "PROCESSO", "PROC", "PROCESSO_APONTADO", "ETAPA", "OPERACAO", "SERVIÇO", "BLOCO", "NUMERO DO BLOCO", "Nº BLOCO", "NUM BLOCO", "N BLOCO", "ID BLOCO", "IDENTIFICAÇÃO", "CÓDIGO"]
        header_row = 6 # fallback padrão
        for i, row in df_full.iterrows():
            row_vals = [str(v).strip().upper() for v in row.values if pd.notna(v)]
            # Se encontrar pelo menos 2 colunas conhecidas na mesma linha, achamos o cabeçalho
            if sum(1 for v in row_vals if v in todos_aliases) >= 2:
                header_row = i
                break
        
        df = pd.read_excel(
            _get_apontamento_file(),
            sheet_name=_get_sheet("SHEET_AP_BD"),
            skiprows=header_row,
            engine="openpyxl"
        )
        # Limpar nomes de colunas
        orig_cols = [str(c).strip().upper() for c in df.columns]
        df.columns = orig_cols

        # Mapeamento robusto
        mapping = {
            "DATA_REG": ["DATA INICIO", "DATA INÍCIO", "DATA DE INÍCIO", "DATA INÍCIO APONTAMENTO", "DATA REG", "DATA", "DATA_REG", "DATA DO APONTAMENTO", "DATA LANÇAMENTO"],
            "MAT_BLOCO": ["MATERIAL+BLOCO", "MAT+BLO", "MATERIAL BLOCO", "MAT/BLO"],
            "NOME_MATERIAL": ["NOME MATERIAL", "MATERIAL", "NOME_MATERIAL", "DESCRIÇÃO MATERIAL"],
            "BLOCO_RAW": ["NUMERO DO BLOCO", "Nº BLOCO", "NUM BLOCO", "BLOCO", "N BLOCO", "ID BLOCO", "IDENTIFICAÇÃO"],
            "PROCESSO_APONTADO": ["PROCESSO", "PROC", "PROCESSO_APONTADO", "ETAPA", "OPERACAO"],
            "SETOR_AP": ["SETOR", "MAQUINA", "MÁQUINA", "SETOR_AP", "LOCAL"],
            "QTD_CH": ["QTD. CHAPAS", "QTD CH (SEM RET & REPASSE)", "QTD CH", "CHAPAS", "QTD_CH", "TOTAL CHAPAS"],
            "QTD_M2": ["QTD M² (SEM RET & REPASSE)", "QTD M²", "QTD M2", "METRAGEM", "QTD M", "QTD_M2", "VOLUME M²"]
        }

        rename_dict = {}
        for target, aliases in mapping.items():
            for alias in aliases:
                if alias.upper() in orig_cols:
                    rename_dict[alias.upper()] = target
                    break
        
        df = df.rename(columns=rename_dict)

        if "DATA_REG" not in df.columns:
            # Tenta encontrar qualquer coluna que tenha 'DATA' no nome se não achou a exata
            for c in df.columns:
                if "DATA" in c.upper():
                    df = df.rename(columns={c: "DATA_REG"})
                    break

        # Conversão robusta de data (DayFirst=True para Brasil)
        df["DATA_REG"] = pd.to_datetime(df["DATA_REG"], errors="coerce", dayfirst=True)
        df = df.dropna(subset=["DATA_REG"])

        df_dia = df[df["DATA_REG"].dt.date == data_alvo_date].copy()

        if df_dia.empty:
            return pd.DataFrame()

        def extrair_bloco(row):
            b = row.get("BLOCO_RAW")
            if pd.isna(b) or str(b).strip() in ["", "nan", "None"]:
                mat_bloco = str(row.get("MAT_BLOCO", ""))
                if "-" in mat_bloco: b = mat_bloco.rsplit("-", 1)[-1]
                else: return ""
            # Limpeza radical: pega só os números antes do ponto e limpa espaços
            return str(b).strip().split(".")[0].upper()

        df_dia["BLOCO"] = df_dia.apply(extrair_bloco, axis=1)

        mapa = get_mapa_resumido_processos()
        def resumir(proc):
            proc_up = str(proc).strip().upper()
            return mapa.get(proc_up, proc_up)

        df_dia["RESUMIDO"] = df_dia["PROCESSO_APONTADO"].apply(resumir)
        df_dia["QTD_CH"] = pd.to_numeric(df_dia["QTD_CH"], errors="coerce").fillna(0)
        df_dia["QTD_M2"] = pd.to_numeric(df_dia.get("QTD_M2", 0), errors="coerce").fillna(0)

        colunas = ["BLOCO", "NOME_MATERIAL", "MAT_BLOCO", "PROCESSO_APONTADO", "RESUMIDO", "SETOR_AP", "QTD_CH", "QTD_M2", "DATA_REG"]
        return df_dia[[c for c in colunas if c in df_dia.columns]].reset_index(drop=True)
    except Exception as e:
        print(f"Erro ao ler apontamentos do dia: {e}")
        return pd.DataFrame()


def get_all_apontamentos():
    """Lê todos os registros da aba de Apontamento (DB)."""
    try:
        file_path = _get_apontamento_file()
        sheet_name = _get_sheet("SHEET_AP_BD")
        
        # Lê as primeiras 20 linhas para encontrar o cabeçalho dinamicamente
        df_scan = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20, engine="openpyxl")
        header_row_idx = 0
        for i, row in df_scan.iterrows():
            row_vals = [str(v).strip().upper() for v in row.values]
            if "PROCESSO" in row_vals or "DATA REG" in row_vals or "MATERIAL+BLOCO" in row_vals:
                header_row_idx = i
                break
        
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx, engine="openpyxl")
        df.columns = [str(c).strip().upper() for c in df.columns]
        return df
    except Exception as e:
        print(f"Erro ao ler todos os apontamentos: {e}")
        return pd.DataFrame()


    """
    Adiciona um novo registro na aba 'BD' do arquivo de Apontamento.
    Busca dinamicamente o cabeçalho e insere na próxima linha vazia.
    """
    try:
        file_path = _get_apontamento_file()
        sheet_name = _get_sheet("SHEET_AP_BD")
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        # Busca dinâmica do cabeçalho (igual ao get)
        header_row_idx = 6 # 1-indexed fallback
        for r in range(1, 21):
            row_vals = [str(ws.cell(row=r, column=c).value).strip().upper() for c in range(1, ws.max_column + 1)]
            if "PROCESSO" in row_vals or "DATA REG" in row_vals or "MATERIAL+BLOCO" in row_vals:
                header_row_idx = r
                break
        
        # Mapeia colunas do cabeçalho
        headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row_idx, column=c).value
            if val:
                headers[str(val).strip().upper()] = c
        
        # Encontra próxima linha vazia (baseada na coluna DATA REG ou PROCESSO)
        col_ref = headers.get("DATA REG") or headers.get("PROCESSO") or 1
        next_row = ws.max_row + 1
        for r in range(header_row_idx + 1, ws.max_row + 2):
            if ws.cell(row=r, column=col_ref).value is None:
                next_row = r
                break
        
        # Mapeamento de campos fixos do sistema para nomes prováveis no Excel
        system_mapping = _get_system_mapping()

        # 1. Primeiro grava os campos mapeados pelo sistema

        # 1. Primeiro grava os campos mapeados pelo sistema
        for key, value in record_dict.items():
            target_col_name = None
            if key in system_mapping:
                # Procura qual alias existe no Excel
                for alias in system_mapping[key]:
                    if alias.upper() in headers:
                        target_col_name = alias.upper()
                        break
            
            # 2. Se não for um campo mapeado, tenta gravar pelo nome exato (para campos dinâmicos como ABRASIVOS, etc)
            if not target_col_name and key.upper() in headers:
                target_col_name = key.upper()

            if target_col_name:
                col_idx = headers[target_col_name]
                ws.cell(row=next_row, column=col_idx, value=value)
        
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Erro ao adicionar apontamento: {e}")
        return False


def _get_system_mapping():
    """Retorna o mapeamento centralizado de campos para colunas Excel."""
    mapping = {
        "DATA_REG": ["DATA REG", "DATA", "DATA_REG"],
        "MAT_BLOCO": ["MATERIAL+BLOCO", "MAT+BLO", "MATERIAL BLOCO"],
        "NOME_MATERIAL": ["NOME MATERIAL", "MATERIAL", "NOME_MATERIAL"],
        "BLOCO_RAW": ["NUMERO DO BLOCO", "Nº BLOCO", "NUM BLOCO", "BLOCO", "N BLOCO", "NUMERO_BLOCO"],
        "PROCESSO_APONTADO": ["PROCESSO", "PROC", "PROCESSO_APONTADO"],
        "SETOR_AP": ["SETOR", "MAQUINA", "MÁQUINA", "SETOR_AP"],
        "QTD_CH": ["QTD. CHAPAS", "QTD CH (SEM RET & REPASSE)", "QTD CH", "CHAPAS", "QTD_CH", "QTD_CHAPAS"],
        "QTD_M2": ["QTD M² (SEM RET & REPASSE)", "QTD M²", "QTD M2", "METRAGEM", "QTD M", "QTDM2"],
        "ESP": ["ESP.", "ESPESSURA", "ESP"],
        "COMP": ["COMP.", "COMPRIMENTO", "COMP"],
        "ALT": ["ALT.", "ALTURA", "ALT"],
        "OPERADOR": ["OPERADOR"],
        "DUREZA": ["DUREZA"],
        "DIA_INICIO": ["DIA INICIO", "DIA_INICIO", "DATA_INICIO"],
        "DIA_FIM": ["DIA FIM", "DIA_FIM", "DATA_FIM"],
        "HORA_INICIO": ["HORA INICIO", "HORA_INICIO"],
        "HORA_FIM": ["HORA FIM", "HORA_FIM"],
        "TEMPO_PROCESSO": ["TEMPO DO PROCESSO", "TEMPO", "TEMPO_PROCESSO"],
        "TURNO": ["TURNO"],
        "TIPO_ACIDO": ["TIPO ACIDO", "TIPO_ACIDO"],
        "TIPO_RESINA": ["TPO RESINA", "TIPO RESINA", "TIPO_RESINA", "TIPO_RES"],
        "QTD_KG": ["QTD.KG", "QTD KG", "QTD_KG", "QTDKG_RES"],
        "TIPO_ENDUR": ["TIPO.ENDUR", "TIPO ENDURECEDOR", "TIPO_ENDUR"],
        "QTD_KG3": ["QTD.KG3", "QTD KG3", "QTD_KG3", "QDKG_ENDUR"],
        "V_24H": ["TEMPO DE SECAGEM", "TEMPO SECAGEM", "24H", "24 H", "V_24H", "TEMPO_SECAGEM"],
        "VEL_ESTEIRA": ["VEL.ESTEIRA", "VEL ESTEIRA", "VEL_ESTEIRA"],
        "VEL_TRAVE": ["VEL.TRAVE", "VEL TRAVE", "VEL_TRAVE"],
        "TIPO_MANTA": ["TIPO MANTA", "TIPO_MANTA"],
        "QTD_MANTA": ["QTD MANTA", "QTD_MANTA"],
        "ID": ["ID", "Nº ID", "ID_APONTAMENTO"]
    }
    # Adiciona mapeamento de abrasivos SAT1 a SAT20
    for i in range(1, 21):
        mapping[f"Seq. Abr. {i}"] = [f"SAT{i}", f"SEQ. ABR. {i}", f"ABRASIVO {i}"]
    return mapping


def get_next_apontamento_id():
    """
    Lê a aba DB do Apontamento e retorna o próximo ID sequencial.
    """
    try:
        file_path = _get_apontamento_file()
        sheet_name = _get_sheet("SHEET_AP_BD")
        
        # Lê apenas a coluna ID se possível, ou as primeiras linhas
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=100)
        
        # Localiza o cabeçalho
        header_row = 6
        for i, row in df.iterrows():
            if "ID" in [str(v).strip().upper() for v in row.values]:
                header_row = i
                break
        
        # Lê a coluna ID a partir do cabeçalho
        df_full = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=header_row, engine="openpyxl")
        df_full.columns = [str(c).strip().upper() for c in df_full.columns]
        
        if "ID" in df_full.columns:
            ids = pd.to_numeric(df_full["ID"], errors="coerce").dropna()
            if not ids.empty:
                return int(ids.max() + 1)
        return 1
    except Exception as e:
        print(f"Erro ao gerar novo ID: {e}")
        return 1


def add_paradas(paradas_list):
    """
    Adiciona registros de paradas na aba 'PARADAS'.
    """
    try:
        file_path = _get_apontamento_file()
        sheet_name = _get_sheet("SHEET_AP_PARADAS")
        wb = openpyxl.load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            # Cria a aba se não existir
            ws = wb.create_sheet(sheet_name)
            headers = ["ID_APONTAMENTO", "MOTIVO", "HORA_INICIO", "HORA_FIM", "TEMPO"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=1, column=c, value=h)
        else:
            ws = wb[sheet_name]

        # Mapeia cabeçalhos existentes
        headers_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                headers_map[str(val).strip().upper()] = c

        next_row = ws.max_row + 1
        for parada in paradas_list:
            for key, val in parada.items():
                col_name = key.upper()
                if col_name in headers_map:
                    ws.cell(row=next_row, column=headers_map[col_name], value=val)
            next_row += 1

        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Erro ao adicionar paradas: {e}")
        return False


def add_apontamento_full(record_dict, paradas_list=None):
    """
    Gera ID, grava apontamento e paradas.
    """
    new_id = get_next_apontamento_id()
    record_dict["ID"] = new_id
    
    if add_apontamento(record_dict):
        if paradas_list:
            for p in paradas_list:
                p["ID_APONTAMENTO"] = new_id
            add_paradas(paradas_list)
        return True, new_id
    return False, None


def add_apontamento_batch(batch_list):
    """
    Grava uma lista de (record_dict, paradas_list) de uma só vez no Excel.
    batch_list: lista de tuplas [(record, paradas), ...]
    """
    try:
        file_path = _get_apontamento_file()
        sheet_bd = _get_sheet("SHEET_AP_BD")
        sheet_paradas = _get_sheet("SHEET_AP_PARADAS")
        
        wb = openpyxl.load_workbook(file_path)
        ws_bd = wb[sheet_bd]
        
        # 1. Prepara cabeçalhos e próxima linha da aba DB
        header_row_idx = 6
        for r in range(1, 21):
            row_vals = [str(ws_bd.cell(row=r, column=c).value).strip().upper() for c in range(1, ws_bd.max_column + 1)]
            if "PROCESSO" in row_vals or "DATA REG" in row_vals or "MATERIAL+BLOCO" in row_vals:
                header_row_idx = r
                break
        
        headers_bd = {}
        for c in range(1, ws_bd.max_column + 1):
            val = ws_bd.cell(row=header_row_idx, column=c).value
            if val:
                headers_bd[str(val).strip().upper()] = c

        system_mapping = _get_system_mapping()

        col_ref = headers_bd.get("DATA REG") or headers_bd.get("PROCESSO") or 1

        col_ref = headers_bd.get("DATA REG") or headers_bd.get("PROCESSO") or 1
        next_row_bd = ws_bd.max_row + 1
        for r in range(header_row_idx + 1, ws_bd.max_row + 2):
            if ws_bd.cell(row=r, column=col_ref).value is None:
                next_row_bd = r
                break

        # 2. Prepara aba PARADAS
        if sheet_paradas not in wb.sheetnames:
            ws_p = wb.create_sheet(sheet_paradas)
            headers_p_list = ["ID_APONTAMENTO", "MOTIVO", "DIA_INICIO", "HORA_INICIO", "DIA_FIM", "HORA_FIM", "TEMPO"]
            for c, h in enumerate(headers_p_list, 1):
                ws_p.cell(row=1, column=c, value=h)
        else:
            ws_p = wb[sheet_paradas]

        headers_p = {}
        for c in range(1, ws_p.max_column + 1):
            val = ws_p.cell(row=1, column=c).value
            if val:
                headers_p[str(val).strip().upper()] = c

        next_row_p = ws_p.max_row + 1

        # Prepara aba INSUMOS
        sheet_insumos_name = _get_sheet("SHEET_AP_INSUMOS")
        if sheet_insumos_name not in wb.sheetnames:
            ws_i = wb.create_sheet(sheet_insumos_name)
            headers_i_list = ["ID_APONTAMENTO", "TIPO_INSUMO", "DESCRICAO", "QUANTIDADE", "UNIDADE", "TEMPO_SECAGEM", "CABECAS", "INSUMO_DETALHE"]
            for c, h in enumerate(headers_i_list, 1):
                ws_i.cell(row=1, column=c, value=h)
        else:
            ws_i = wb[sheet_insumos_name]

        headers_i = {}
        for c in range(1, ws_i.max_column + 1):
            val = ws_i.cell(row=1, column=c).value
            if val:
                headers_i[str(val).strip().upper()] = c

        next_row_i = ws_i.max_row + 1

        # 3. Itera sobre o batch e grava
        for record, paradas, insumos in batch_list:
            # Grava record no DB
            for key, value in record.items():
                target_col = None
                if key in system_mapping:
                    for alias in system_mapping[key]:
                        if alias.upper() in headers_bd:
                            target_col = alias.upper()
                            break
                if not target_col and key.upper() in headers_bd:
                    target_col = key.upper()
                
                if target_col:
                    ws_bd.cell(row=next_row_bd, column=headers_bd[target_col], value=value)
            
            # Grava paradas
            if paradas:
                for p in paradas:
                    p["ID_APONTAMENTO"] = record.get("ID")
                    for key, value in p.items():
                        target_col = None
                        if key in system_mapping:
                            for alias in system_mapping[key]:
                                if alias.upper() in headers_p:
                                    target_col = alias.upper()
                                    break
                        if not target_col and key.upper() in headers_p:
                            target_col = key.upper()
                        
                        if target_col:
                            ws_p.cell(row=next_row_p, column=headers_p[target_col], value=value)
                    next_row_p += 1

            # Grava insumos
            if insumos:
                for i_data in insumos:
                    i_data["ID_APONTAMENTO"] = record.get("ID")
                    for key, value in i_data.items():
                        target_col = None
                        if key in system_mapping:
                            for alias in system_mapping[key]:
                                if alias.upper() in headers_i:
                                    target_col = alias.upper()
                                    break
                        if not target_col and key.upper() in headers_i:
                            target_col = key.upper()
                        
                        if target_col:
                            ws_i.cell(row=next_row_i, column=headers_i[target_col], value=value)
                    next_row_i += 1
            
            next_row_bd += 1

        # 4. Tenta expandir Tabela do Excel (ListObject) se existir
        try:
            if ws_bd.tables:
                for table_name, table in ws_bd.tables.items():
                    current_ref = table.ref # ex: "A1:T18881"
                    parts = current_ref.split(":")
                    if len(parts) == 2:
                        start_cell = parts[0]
                        import re
                        col_part = re.sub(r'\d+', '', parts[1])
                        new_ref = f"{start_cell}:{col_part}{next_row_bd - 1}"
                        table.ref = new_ref
        except: pass

        # 5. Copia estilos da linha anterior para as novas linhas
        try:
            prev_row_idx = next_row_bd - len(batch_list) - 1
            if prev_row_idx > header_row_idx:
                from copy import copy
                for r_idx in range(next_row_bd - len(batch_list), next_row_bd):
                    for c_idx in range(1, ws_bd.max_column + 1):
                        source_cell = ws_bd.cell(row=prev_row_idx, column=c_idx)
                        target_cell = ws_bd.cell(row=r_idx, column=c_idx)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = copy(source_cell.number_format)
                            target_cell.protection = copy(source_cell.protection)
                            target_cell.alignment = copy(source_cell.alignment)
        except: pass

        wb.save(file_path)
        return True, f"{len(batch_list)} registros salvos."
    except Exception as e:
        import traceback
        error_msg = f"Erro no salvamento em lote: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        return False, str(e)


def get_apontamentos_por_bloco(bloco_id):
    """
    Busca o histórico de processos já realizados para um bloco específico.
    """
    try:
        file_path = _get_apontamento_file()
        sheet_name = _get_sheet("SHEET_AP_BD")
        
        # Lê as primeiras 20 linhas para encontrar o cabeçalho dinamicamente
        df_scan = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20, engine="openpyxl")
        header_row_idx = 0
        for i, row in df_scan.iterrows():
            row_vals = [str(v).strip().upper() for v in row.values]
            if "PROCESSO" in row_vals or "DATA REG" in row_vals or "DATA_REG" in row_vals or "NUMERO_BLOCO" in row_vals:
                header_row_idx = i
                break
        
        # Agora lê os dados de verdade a partir do cabeçalho encontrado
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx, engine="openpyxl")
        df.columns = [str(c).strip().upper() for c in df.columns]
        
        # Mapeia qual coluna é o bloco usando o sistema de mapping robusto
        mapping = _get_system_mapping()
        col_bloco = None
        for alias in mapping.get("BLOCO_RAW", []):
            if alias.upper() in df.columns:
                col_bloco = alias.upper()
                break
        
        if not col_bloco:
            return pd.DataFrame()
            
        # Filtra o bloco
        df[col_bloco] = df[col_bloco].astype(str).str.strip().str.split(".").str[0].str.upper()
        historico = df[df[col_bloco] == str(bloco_id).strip().upper()].copy()
        
        # Seleciona colunas úteis para o histórico (mapeando para o que existir no REV 2)
        res_cols = []
        wanted = ["DATA_REG", "PROCESSO_APONTADO", "SETOR_AP", "QTD_CH", "QTD_M2", "OPERADOR", "HORA_INICIO", "HORA_FIM"]
        for w in wanted:
            for alias in mapping.get(w, []):
                if alias.upper() in historico.columns:
                    res_cols.append(alias.upper())
                    break
        
        if not res_cols: return pd.DataFrame()
        return historico[res_cols].sort_values(by=res_cols[0], ascending=False)
        
    except Exception as e:
        print(f"Erro ao buscar histórico do bloco: {e}")
        return pd.DataFrame()


def get_bloco_info(bloco_id):
    """
    Busca informações de material e medidas na PLANILHA BLOCOS.xlsb.
    Retorna dict com MATERIAL, COMP_LIQUIDO, ALT_LIQUIDO, LARG_LIQUIDO ou None.
    """
    try:
        cfg = get_config()
        path = cfg.get("BLOCKS_FILE")
        if not path or not os.path.exists(path):
            return None

        # Lê apenas as colunas necessárias para performance
        # Header está na linha 9 (skiprows=8)
        df = pd.read_excel(path, engine="pyxlsb", sheet_name="PLAN. BLOCOS", skiprows=8)
        
        # Normaliza nomes de colunas
        df.columns = [str(c).strip().upper() for c in df.columns]
        
        # Busca o bloco (converte para string e remove .0)
        bloco_busca = str(bloco_id).strip().split(".")[0].upper()
        df["N_BLOCO_STR"] = df["N_BLOCO"].astype(str).str.strip().str.split(".").str[0].str.upper()
        
        match = df[df["N_BLOCO_STR"] == bloco_busca]
        if not match.empty:
            row = match.iloc[0]
            return {
                "MATERIAL": str(row.get("MATERIAL", "")).strip().upper(),
                "COMP": float(row.get("COMP_LIQUIDO", 0)),
                "ALT": float(row.get("ALT_LIQUIDO", 0)),
                "LARG": float(row.get("LARG_LIQUIDO", 0)),
                "SOURCE": "Planilha de Blocos"
            }
    except Exception as e:
        print(f"Erro ao buscar na PLANILHA BLOCOS: {e}")

    # --- FONTE 2: ESTOQUE DE CHAPAS (ENTRADAS) ---
    try:
        cfg = get_config()
        path = cfg.get("CHAPAS_FILE")
        if path and os.path.exists(path):
            df_ch = pd.read_excel(path, sheet_name="ENTRADAS", engine="openpyxl")
            # Normaliza nomes de colunas
            df_ch.columns = [str(c).strip().upper() for c in df_ch.columns]
            
            # Mapeamento flexível de colunas
            col_bloco = next((c for c in ["BLOCO", "Nº BLOCO", "N_BLOCO", "NUMERO DO BLOCO", "MAT+BLO"] if c in df_ch.columns), None)
            col_mat = next((c for c in ["MATERIAL", "NOME MATERIAL", "NOME_MATERIAL"] if c in df_ch.columns), None)
            col_comp = next((c for c in ["COMP", "COMP.", "COMPRIMENTO", "C", "COMP_LIQUIDO"] if c in df_ch.columns), None)
            col_alt = next((c for c in ["ALT", "ALT.", "ALTURA", "A", "ALT_LIQUIDO", "LARGURA"] if c in df_ch.columns), None)
            col_esp = next((c for c in ["ESP", "ESP.", "ESPESSURA", "E", "LARG_LIQUIDO"] if c in df_ch.columns), None)

            if col_bloco:
                df_ch["N_BLOCO_STR"] = df_ch[col_bloco].astype(str).str.strip().str.split(".").str[0].str.upper()
                bloco_busca = str(bloco_id).strip().split(".")[0].upper()
                
                match = df_ch[df_ch["N_BLOCO_STR"] == bloco_busca]
                if not match.empty:
                    row = match.iloc[0]
                    return {
                        "MATERIAL": str(row.get(col_mat, "")).strip().upper() if col_mat else "",
                        "COMP": float(pd.to_numeric(row.get(col_comp, 0), errors="coerce") or 0) if col_comp else 0,
                        "ALT": float(pd.to_numeric(row.get(col_alt, 0), errors="coerce") or 0) if col_alt else 0,
                        "LARG": float(pd.to_numeric(row.get(col_esp, 0), errors="coerce") or 0) if col_esp else 0,
                        "SOURCE": "Estoque de Chapas"
                    }
    except Exception as e:
        print(f"Erro ao buscar no ESTOQUE CHAPAS: {e}")

    return None


def get_listas_endurentes():
    """Retorna a lista de endurecedores e suas proporções da aba LISTAS."""
    try:
        file_path = _get_apontamento_file()
        df = pd.read_excel(file_path, sheet_name='LISTAS', engine="openpyxl")
        df.columns = [str(c).strip().upper() for c in df.columns]
        return df.to_dict(orient='records')
    except Exception as e:
        print(f"Erro ao carregar lista de endurecedores: {e}")
        return []
