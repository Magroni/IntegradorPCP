import openpyxl
import os

def correct():
    file_path = 'Apontamento Produção (REV 2).xlsx'
    if not os.path.exists(file_path):
        print(f"Erro: Arquivo {file_path} nao existe.")
        return
        
    wb = openpyxl.load_workbook(file_path)
    sheet_name = 'DB'
    ws = wb[sheet_name]
    
    # 1. Encontra a linha com cabeçalhos para mapear as colunas
    header_row_idx = 1
    for r in range(1, 21):
        row_vals = [str(ws.cell(row=r, column=c).value).strip().upper() for c in range(1, ws.max_column + 1)]
        if "PROCESSO" in row_vals or "DATA REG" in row_vals or "NUMERO_BLOCO" in row_vals:
            header_row_idx = r
            break
            
    headers = {str(ws.cell(row=header_row_idx, column=c).value).strip().upper(): c for c in range(1, ws.max_column + 1)}
    print("Mapeamento de colunas:", headers)
    
    col_id = headers.get("ID")
    col_setor = headers.get("SETOR")
    
    if not col_id or not col_setor:
        print("Erro: Colunas ID ou SETOR nao encontradas.")
        return
        
    # 2. Busca a linha onde o ID é 18995
    row_found = None
    for r in range(header_row_idx + 1, ws.max_row + 1):
        cell_id = ws.cell(row=r, column=col_id).value
        try:
            if int(cell_id) == 18995:
                row_found = r
                break
        except:
            continue
            
    if not row_found:
        print("Erro: Registro com ID 18995 nao encontrado.")
        return
        
    # 3. Atualiza o SETOR para CIMEF
    old_setor = ws.cell(row=row_found, column=col_setor).value
    ws.cell(row=row_found, column=col_setor, value="CIMEF")
    
    # 4. Salva o arquivo
    wb.save(file_path)
    print(f"Sucesso! Registro ID 18995 (linha {row_found}) atualizado: SETOR alterado de '{old_setor}' para 'CIMEF'.")

if __name__ == "__main__":
    correct()
