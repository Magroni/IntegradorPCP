import shutil
import openpyxl
import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import data_manager as dm

# Define test delete function
def delete_apontamento_test(file_path, id_apontamento):
    wb = openpyxl.load_workbook(file_path)
    id_val_float = float(id_apontamento)
    
    # 1. DB Sheet
    sheet_bd = "DB"
    ws_bd = wb[sheet_bd]
    header_row_idx = 0
    for r in range(1, 21):
        row_vals = [str(ws_bd.cell(row=r, column=c).value).strip().upper() for c in range(1, ws_bd.max_column + 1)]
        if "PROCESSO" in row_vals or "DATA REG" in row_vals or "MATERIAL+BLOCO" in row_vals:
            header_row_idx = r
            break
    col_id_idx = None
    for c in range(1, ws_bd.max_column + 1):
        val = ws_bd.cell(row=header_row_idx, column=c).value
        if val and str(val).strip().upper() == "ID":
            col_id_idx = c
            break
    if col_id_idx:
        for r in range(ws_bd.max_row, header_row_idx, -1):
            cell_val = ws_bd.cell(row=r, column=col_id_idx).value
            try:
                if cell_val is not None and float(cell_val) == id_val_float:
                    print(f"Deleting row {r} in DB")
                    ws_bd.delete_rows(r, 1)
            except:
                pass

    # 2. PARADAS
    sheet_paradas = "PARADAS"
    ws_p = wb[sheet_paradas]
    col_id_idx = None
    for c in range(1, ws_p.max_column + 1):
        val = ws_p.cell(row=1, column=c).value
        if val and str(val).strip().upper() == "ID_APONTAMENTO":
            col_id_idx = c
            break
    if col_id_idx:
        for r in range(ws_p.max_row, 1, -1):
            cell_val = ws_p.cell(row=r, column=col_id_idx).value
            try:
                if cell_val is not None and float(cell_val) == id_val_float:
                    print(f"Deleting row {r} in PARADAS")
                    ws_p.delete_rows(r, 1)
            except:
                pass

    # 3. INSUMOS
    sheet_insumos = "INSUMOS"
    ws_i = wb[sheet_insumos]
    col_id_idx = None
    for c in range(1, ws_i.max_column + 1):
        val = ws_i.cell(row=1, column=c).value
        if val and str(val).strip().upper() == "ID_APONTAMENTO":
            col_id_idx = c
            break
    if col_id_idx:
        for r in range(ws_i.max_row, 1, -1):
            cell_val = ws_i.cell(row=r, column=col_id_idx).value
            try:
                if cell_val is not None and float(cell_val) == id_val_float:
                    print(f"Deleting row {r} in INSUMOS")
                    ws_i.delete_rows(r, 1)
            except:
                pass

    wb.save(file_path)
    print("Test file saved successfully!")

# Copy original file to test file
orig_file = dm._get_apontamento_file()
test_file = os.path.join(os.path.dirname(__file__), "test_apontamento.xlsx")
shutil.copyfile(orig_file, test_file)
print(f"Copied {orig_file} to {test_file}")

# Verify current rows in test file for ID 18883
import pandas as pd
db_df = pd.read_excel(test_file, sheet_name="DB")
db_df.columns = [str(c).strip().upper() for c in db_df.columns]
print("Rows in DB before delete matching 18883:", len(db_df[db_df["ID"] == 18883]))

par_df = pd.read_excel(test_file, sheet_name="PARADAS")
par_df.columns = [str(c).strip().upper() for c in par_df.columns]
print("Rows in PARADAS before delete matching 18883:", len(par_df[par_df["ID_APONTAMENTO"] == 18883]))

# Delete 18883
delete_apontamento_test(test_file, 18883)

# Verify rows after delete
db_df2 = pd.read_excel(test_file, sheet_name="DB")
db_df2.columns = [str(c).strip().upper() for c in db_df2.columns]
print("Rows in DB after delete matching 18883:", len(db_df2[db_df2["ID"] == 18883]))

par_df2 = pd.read_excel(test_file, sheet_name="PARADAS")
par_df2.columns = [str(c).strip().upper() for c in par_df2.columns]
print("Rows in PARADAS after delete matching 18883:", len(par_df2[par_df2["ID_APONTAMENTO"] == 18883]))

# Clean up test file
os.remove(test_file)
print("Test file removed.")
