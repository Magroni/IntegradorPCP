# -*- coding: utf-8 -*-
import pandas as pd
import os
import openpyxl

BASE_DIR = r"z:\PCP\PROJETOS MARLON\ProgramarProd"
file1 = os.path.join(BASE_DIR, "Apontamento Produção (REV 1).xlsx")
file2 = os.path.join(BASE_DIR, "Apontamento Produção (REV 2).xlsx")

print("=== DEEP SEARCH FOR PARADAS ===")

# Search in REV 1
print("\n--- 1. Searching in REV 1 (Apontamento Produção (REV 1).xlsx) ---")
try:
    # A. Search in "BD" sheet
    print("Reading REV 1 'BD' sheet...")
    df_bd = pd.read_excel(file1, sheet_name="BD", skiprows=6, engine="openpyxl")
    df_bd.columns = [str(c).strip().upper() for c in df_bd.columns]
    
    # Search for checklist, lixamento, levantamento in all string columns of df_bd
    search_terms = ["CHECK", "LIXA", "LEVANTA"]
    print("Searching for keywords:", search_terms)
    
    for term in search_terms:
        # Check in MATERIAL+BLOCO
        matches_mat = df_bd[df_bd["MATERIAL+BLOCO"].astype(str).str.upper().str.contains(term, na=False)]
        # Check in PROCESSO
        matches_proc = df_bd[df_bd["PROCESSO"].astype(str).str.upper().str.contains(term, na=False)]
        
        print(f"\nKeyword '{term}':")
        print(f"  Matches in MATERIAL+BLOCO: {len(matches_mat)}")
        if len(matches_mat) > 0:
            print("  Samples from MATERIAL+BLOCO:")
            print(matches_mat[["DATA REG", "MATERIAL+BLOCO", "PROCESSO", "SETOR"]].head(5).to_string())
            
        print(f"  Matches in PROCESSO: {len(matches_proc)}")
        if len(matches_proc) > 0:
            print("  Samples from PROCESSO:")
            print(matches_proc[["DATA REG", "MATERIAL+BLOCO", "PROCESSO", "SETOR"]].head(5).to_string())

    # B. Search in other sheets in REV 1 related to Paradas
    print("\nReading other sheets in REV 1...")
    wb1 = openpyxl.load_workbook(file1, read_only=True)
    for sheet in wb1.sheetnames:
        if "PARADA" in sheet.upper() or "PERDA" in sheet.upper() or "HIST" in sheet.upper():
            print(f"  Searching in sheet '{sheet}'...")
            try:
                df_s = pd.read_excel(file1, sheet_name=sheet, nrows=500, engine="openpyxl")
                # Search for checklist, lixamento, levantamento in all cell values
                for term in search_terms:
                    mask = df_s.astype(str).apply(lambda x: x.str.upper().str.contains(term)).any(axis=1)
                    matches = df_s[mask]
                    if len(matches) > 0:
                        print(f"    Found {len(matches)} matches for '{term}' in sheet '{sheet}'!")
                        print("    Columns:", df_s.columns.tolist()[:8])
                        print(matches.head(3).to_string())
            except Exception as e:
                print(f"    Error reading sheet {sheet}: {e}")
    wb1.close()

except Exception as e:
    print("Error searching in REV 1:", e)

# Search in REV 2
print("\n--- 2. Searching in REV 2 (Apontamento Produção (REV 2).xlsx) ---")
try:
    wb2 = openpyxl.load_workbook(file2, read_only=True)
    print("REV 2 Sheet names:", wb2.sheetnames)
    wb2.close()
    
    # Search in DB, PARADAS, INSUMOS, HISTÓRICO PARADAS
    for sheet in ["DB", "PARADAS", "HISTÓRICO PARADAS"]:
        try:
            print(f"  Reading sheet '{sheet}' from REV 2...")
            df_s = pd.read_excel(file2, sheet_name=sheet, engine="openpyxl")
            for term in ["CHECK", "LIXA", "LEVANTA"]:
                mask = df_s.astype(str).apply(lambda x: x.str.upper().str.contains(term)).any(axis=1)
                matches = df_s[mask]
                if len(matches) > 0:
                    print(f"    Found {len(matches)} matches for '{term}' in sheet '{sheet}'!")
                    print(matches.head(3).to_string())
        except Exception as e:
            print(f"  Error reading sheet {sheet} from REV 2: {e}")
            
except Exception as e:
    print("Error searching in REV 2:", e)
