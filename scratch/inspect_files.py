# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd
import os

BASE_DIR = r"z:\PCP\PROJETOS MARLON\ProgramarProd"
file1 = os.path.join(BASE_DIR, "Apontamento Produção (REV 1).xlsx")
file2 = os.path.join(BASE_DIR, "Apontamento Produção (REV 2).xlsx")

print("=== INSPECTING FILES ===")

# File 1 (REV 1)
print(f"\n1. File: {os.path.basename(file1)}")
try:
    wb1 = openpyxl.load_workbook(file1, read_only=True)
    print("Sheets in REV 1:", wb1.sheetnames)
    wb1.close()
    
    # Read headers
    df1 = pd.read_excel(file1, sheet_name="BD", skiprows=6, nrows=3, engine="openpyxl")
    print("REV 1 Columns:")
    for idx, col in enumerate(df1.columns):
        print(f"  [{idx}] {col}")
    print("\nREV 1 First row sample:")
    if not df1.empty:
        print(df1.iloc[0].to_dict())
    else:
        print("No rows found!")
except Exception as e:
    print("Error reading REV 1:", e)

# File 2 (REV 2)
print(f"\n2. File: {os.path.basename(file2)}")
try:
    wb2 = openpyxl.load_workbook(file2, read_only=True)
    print("Sheets in REV 2:", wb2.sheetnames)
    wb2.close()
    
    sheet_name2 = "DB" if "DB" in wb2.sheetnames else wb2.sheetnames[0]
    df2 = pd.read_excel(file2, sheet_name=sheet_name2, nrows=3, engine="openpyxl")
    print(f"REV 2 Columns (sheet: {sheet_name2}):")
    for idx, col in enumerate(df2.columns):
        print(f"  [{idx}] {col}")
    print("\nREV 2 First row sample:")
    if not df2.empty:
        print(df2.iloc[0].to_dict())
    else:
        print("No rows found!")
except Exception as e:
    print("Error reading REV 2:", e)
